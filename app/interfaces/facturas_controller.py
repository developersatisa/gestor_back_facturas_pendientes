from fastapi import APIRouter, Depends, Query, HTTPException, status, Response
from typing import List, Optional, Dict, Any
from datetime import date, datetime
from sqlalchemy.orm import Session
from sqlalchemy import text
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.application.obtener_facturas_filtradas import ObtenerFacturasFiltradas
from app.application.obtener_estadisticas_facturas import ObtenerEstadisticasFacturas
from app.application.obtener_facturas_agrupadas_por_cliente import ObtenerFacturasAgrupadasPorCliente
from app.infrastructure.repositorio_facturas_simple import RepositorioFacturas, RepositorioClientes
from app.infrastructure.data_enrichers import enrich_factura_with_cliente
from app.utils.cliente_helpers import build_cliente_cache_key, format_cliente_data
from app.utils.factura_helpers import format_factura_search_result
from app.utils.error_handlers import handle_error
from app.domain.constants import SOCIEDADES_NAMES
from app.infrastructure.repositorio_gestion import RepositorioGestion
from app.infrastructure.repositorio_registro_facturas import RepositorioRegistroFacturas
from app.domain.models.Factura import Factura
from app.config.database import get_facturas_db, get_clientes_db, get_gestion_db
import logging

router = APIRouter()

# Configuración de logging
logger = logging.getLogger("facturas")

# Dependency para el repositorio
def get_repo_facturas(db: Session = Depends(get_facturas_db)):
    return RepositorioFacturas(db)

def get_repo_clientes(db: Session = Depends(get_clientes_db)):
    return RepositorioClientes(db)
 
def get_repo_gestion(db: Session = Depends(get_gestion_db)):
    return RepositorioGestion(db)

@router.get("/api/facturas-cliente/{idcliente}", response_model=List[dict], tags=["Facturas"])
def obtener_facturas_cliente(
    idcliente: str,
    sociedad: Optional[str] = Query(None, description="Filtro por código de sociedad (CPY_0)"),
    repo_facturas: RepositorioFacturas = Depends(get_repo_facturas),
    repo_clientes: RepositorioClientes = Depends(get_repo_clientes),
    repo_gestion: RepositorioGestion = Depends(get_repo_gestion),
):
    try:
        # Obtener facturas del cliente específico
        use_case = ObtenerFacturasFiltradas(repo_facturas)
        facturas = use_case.execute(
            sociedad=sociedad,
            tercero=idcliente,  # Usar el idcliente como tercero
        )
        
        # Agregar datos del cliente a cada factura
        facturas_con_cliente = [
            enrich_factura_with_cliente(f.copy(), repo_clientes)
            for f in facturas
        ]
        
        return facturas_con_cliente
    except Exception as e:
        logger.error(f"Error al obtener facturas del cliente {idcliente}: {e}")
        raise handle_error(e, f"obtener facturas del cliente {idcliente}", "Error interno del servidor")

@router.get("/api/facturas/buscar", response_model=List[dict], tags=["Facturas"])
def buscar_factura_por_numero(
    numero: str = Query(..., min_length=2, description="Coincidencia con ACCNUM_0 o NUM_0"),
    repo_facturas: RepositorioFacturas = Depends(get_repo_facturas),
    repo_clientes: RepositorioClientes = Depends(get_repo_clientes),
):
    try:
        patron = numero.strip()
        if not patron:
            return []

        coincidencias = repo_facturas.buscar_por_numero(patron)
        if not coincidencias:
            coincidencias_todas = repo_facturas.buscar_por_numero_incluyendo_pagadas(patron)
            if coincidencias_todas:
                factura_ref = coincidencias_todas[0]
                nombre_factura = (
                    factura_ref.get("nombre_factura")
                    or f"{factura_ref.get('tipo', 'N/D')}-{factura_ref.get('asiento', 'N/D')}"
                )
                tercero = factura_ref.get("tercero") or "N/D"
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=(
                        f"La factura {nombre_factura} del tercero {tercero} no tiene saldo pendiente. "
                        "Revisa la pestaña de facturas pagadas o el historial."
                    ),
                )
            return []

        cache_clientes: Dict[str, Optional[Dict[str, Any]]] = {}
        respuesta: List[Dict[str, Any]] = []

        for item in coincidencias:
            tercero = str(item.get("tercero") or "").strip()
            datos_cliente = None
            
            if tercero:
                cache_key = build_cliente_cache_key(tercero)
                if cache_key not in cache_clientes:
                    cache_clientes[cache_key] = repo_clientes.obtener_cliente(tercero)
                datos_cliente = format_cliente_data(cache_clientes.get(cache_key))

            respuesta.append(format_factura_search_result(item, datos_cliente))

        return respuesta
    except Exception as exc:
        logger.error("Error buscando factura '%s': %s", numero, exc)
        raise handle_error(exc, f"buscar factura '{numero}'", "No se pudo buscar la factura")

@router.get("/api/clientes-con-resumen", response_model=List[dict], tags=["Clientes"])
def obtener_clientes_con_resumen(
    sociedad: Optional[str] = Query(None, description="Filtro por código de sociedad (CPY_0)"),
    tercero: Optional[str] = Query(None),
    fecha_desde: Optional[date] = Query(None),
    fecha_hasta: Optional[date] = Query(None),
    nivel_reclamacion: Optional[int] = Query(None),
    repo_facturas: RepositorioFacturas = Depends(get_repo_facturas),
    repo_clientes: RepositorioClientes = Depends(get_repo_clientes),
    repo_gestion: RepositorioGestion = Depends(get_repo_gestion),
):
    try:
        # Obtener facturas agrupadas por cliente
        use_case = ObtenerFacturasAgrupadasPorCliente(repo_facturas, repo_clientes, repo_gestion)
        try:
            clientes_con_facturas = use_case.execute(
                sociedad=sociedad,
                tercero=tercero,
                fecha_desde=fecha_desde,
                fecha_hasta=fecha_hasta,
                nivel_reclamacion=nivel_reclamacion,
            )
        except Exception as inner_err:
            # No tumbar el endpoint: devolver lista vacía y loguear el detalle
            logger.error(f"Error en use_case clientes-con-resumen: {inner_err}")
            clientes_con_facturas = []
        
        return clientes_con_facturas
    except Exception as e:
        logger.error(f"Error al obtener clientes con resumen: {e}")
        # Fallback final seguro: no romper el cliente si algo inesperado sucede
        return []

@router.get("/api/estadisticas", response_model=dict, tags=["Estadísticas"])
def obtener_estadisticas(
    repo_facturas: RepositorioFacturas = Depends(get_repo_facturas),
    repo_clientes: RepositorioClientes = Depends(get_repo_clientes),
):
    try:
        use_case = ObtenerEstadisticasFacturas(repo_facturas, repo_clientes)
        estadisticas = use_case.execute()
        return estadisticas
    except Exception as e:
        logger.error(f"Error al obtener estadísticas: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"{e}")


@router.get("/api/estadisticas/excel", tags=["Estadísticas"])
def descargar_excel_empresas_por_sociedad(
    repo_facturas: RepositorioFacturas = Depends(get_repo_facturas),
    repo_clientes: RepositorioClientes = Depends(get_repo_clientes),
):
    """Genera un Excel con las empresas y sus deudas agrupadas por sociedad (GRUPO ATISA, SELIER, ASESORES TITULADOS)"""
    try:
        # Verificar que el motor sea MSSQL
        try:
            bind = repo_facturas.db.get_bind()  # type: ignore[attr-defined]
            if not bind or bind.dialect.name != 'mssql':
                raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail="Este endpoint requiere conexión MSSQL")
        except Exception:
            raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail="Error detectando motor de BD")
        
        # Consulta SQL para validar totales por sociedad (igual que dashboard)
        query_totales_sociedades = """
        WITH base AS (
          SELECT CPY_0, AMTCUR_0, PAYCUR_0, SNS_0, FLGCLE_0, DUDDAT_0, TYP_0
          FROM x3v12.ATISAINT.GACCDUDATE
          WHERE SAC_0 IN ('4300','4302') AND TYP_0 NOT IN ('AA','ZZ')
            AND DUDDAT_0 < GETDATE() AND FLGCLE_0 <> 2
        )
        SELECT
          CPY_0 as sociedad,
          SUM(CASE
                WHEN (SNS_0=-1 OR FLGCLE_0=-1 OR AMTCUR_0<0) THEN -ABS(AMTCUR_0)
                WHEN (FLGCLE_0=1) AND (SNS_0<>-1) AND (AMTCUR_0-ISNULL(PAYCUR_0,0))>0 THEN (AMTCUR_0-ISNULL(PAYCUR_0,0))
                ELSE 0
              END) as monto_total
        FROM base
        WHERE CPY_0 IN ('S005','S001','S010')
        GROUP BY CPY_0
        HAVING SUM(CASE
                      WHEN (SNS_0=-1 OR FLGCLE_0=-1 OR AMTCUR_0<0) THEN -ABS(AMTCUR_0)
                      WHEN (FLGCLE_0=1) AND (SNS_0<>-1) AND (AMTCUR_0-ISNULL(PAYCUR_0,0))>0 THEN (AMTCUR_0-ISNULL(PAYCUR_0,0))
                      ELSE 0
                    END) > 0
        ORDER BY CPY_0;
        """
        
        # Consulta SQL para obtener facturas individuales por empresa y sociedad
        query_facturas = """
        SELECT 
          BPR_0 as tercero,
          CPY_0 as sociedad,
          TYP_0 as tipo,
          ACCNUM_0 as asiento,
          NUM_0 as nombre_factura,
          DUDDAT_0 as vencimiento,
          CASE
            WHEN (SNS_0=-1 OR FLGCLE_0=-1 OR AMTCUR_0<0) THEN -ABS(AMTCUR_0)
            WHEN (FLGCLE_0=1) AND (SNS_0<>-1) AND (AMTCUR_0-ISNULL(PAYCUR_0,0))>0 THEN (AMTCUR_0-ISNULL(PAYCUR_0,0))
            ELSE 0
          END as monto_pendiente
        FROM x3v12.ATISAINT.GACCDUDATE
        WHERE SAC_0 IN ('4300','4302') 
          AND TYP_0 NOT IN ('AA','ZZ')
          AND DUDDAT_0 < GETDATE() 
          AND FLGCLE_0 <> 2
          AND CPY_0 IN ('S005','S001','S010')
          AND (
            (SNS_0=-1 OR FLGCLE_0=-1 OR AMTCUR_0<0) OR
            ((FLGCLE_0=1) AND (SNS_0<>-1) AND (AMTCUR_0-ISNULL(PAYCUR_0,0))>0)
          )
        ORDER BY sociedad, tercero, monto_pendiente DESC;
        """
        
        # Validar totales con el query del dashboard
        result_totales = repo_facturas.db.execute(text(query_totales_sociedades))
        totales_validacion = {}
        for row in result_totales:
            totales_validacion[row.sociedad] = float(row.monto_total)
        
        # Ejecutar consulta de facturas individuales
        result = repo_facturas.db.execute(text(query_facturas))
        facturas_data = result.fetchall()
        
        from app.domain.constants import SOCIEDADES_NAMES
        
        # Agrupar facturas por sociedad y empresa
        # Estructura: {sociedad: [{tercero, nombre, facturas: [{tipo, asiento, nombre, vencimiento, monto}], total}]}
        datos_por_sociedad = {}
        terceros_procesados = {}  # Cache de nombres de empresas
        
        for row in facturas_data:
            sociedad = row.sociedad
            tercero = str(row.tercero)
            tipo = str(row.tipo) if row.tipo else ''
            asiento = str(row.asiento) if row.asiento else ''
            nombre_factura = str(row.nombre_factura) if row.nombre_factura else f'{tipo}-{asiento}'
            vencimiento = row.vencimiento
            monto = float(row.monto_pendiente) if row.monto_pendiente else 0.0
            
            # Inicializar estructura si no existe
            if sociedad not in datos_por_sociedad:
                datos_por_sociedad[sociedad] = {}
            
            if tercero not in datos_por_sociedad[sociedad]:
                # Obtener nombre de la empresa (usar cache si está disponible)
                tercero_key = tercero.strip()
                if tercero_key not in terceros_procesados:
                    try:
                        tercero_sin_ceros = str(int(tercero))
                    except Exception:
                        tercero_sin_ceros = tercero
                    datos_cliente = repo_clientes.obtener_cliente(tercero_sin_ceros)
                    nombre_empresa = datos_cliente.get('razsoc', 'Sin nombre') if datos_cliente else 'Sin nombre'
                    terceros_procesados[tercero_key] = nombre_empresa
                else:
                    nombre_empresa = terceros_procesados[tercero_key]
                
                datos_por_sociedad[sociedad][tercero] = {
                    'tercero': tercero,
                    'nombre': nombre_empresa,
                    'facturas': [],
                    'total': 0.0
                }
            
            # Agregar factura
            datos_por_sociedad[sociedad][tercero]['facturas'].append({
                'tipo': tipo,
                'asiento': asiento,
                'nombre': nombre_factura,
                'vencimiento': vencimiento,
                'monto': monto
            })
            datos_por_sociedad[sociedad][tercero]['total'] += monto
        
        # Convertir a lista y ordenar por total descendente dentro de cada sociedad
        # IMPORTANTE: Filtrar empresas con saldo neto <= 0 (igual que el listado de empresas)
        for sociedad in datos_por_sociedad:
            empresas_list = []
            for tercero, datos in datos_por_sociedad[sociedad].items():
                # Solo incluir empresas con saldo neto > 0 (igual que el listado de empresas)
                if datos['total'] > 0:
                    # Ordenar facturas por monto descendente
                    datos['facturas'].sort(key=lambda x: x['monto'], reverse=True)
                    empresas_list.append(datos)
            empresas_list.sort(key=lambda x: x['total'], reverse=True)
            datos_por_sociedad[sociedad] = empresas_list
        
        # Crear el Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empresas por Sociedad"
        
        # Estilos
        header_fill = PatternFill(start_color="006341", end_color="006341", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        
        # Encabezado principal (ahora hasta K para incluir espacios)
        ws.merge_cells('A1:K1')
        ws['A1'] = 'Informe de Deudas por Empresa y Sociedad'
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        
        ws['A2'] = f'Fecha de generación: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'
        ws['A2'].font = Font(size=10, italic=True)
        
        # Obtener datos de las tres sociedades
        sociedades_orden = ['S005', 'S010', 'S001']  # Grupo Atisa, Selier, Asesores Titulados
        datos_sociedades = {}
        for soc_codigo in sociedades_orden:
            if soc_codigo in datos_por_sociedad:
                datos_sociedades[soc_codigo] = datos_por_sociedad[soc_codigo]
            else:
                datos_sociedades[soc_codigo] = []  # Asegurar que todas las sociedades existan
        
        # Encontrar el máximo de empresas para saber cuántas filas necesitamos
        max_empresas = max([len(datos_sociedades.get(soc, [])) for soc in sociedades_orden], default=0)
        
        row = 4
        # Columnas con espacios entre sociedades: Grupo Atisa (A-C), espacio (D), Selier (E-G), espacio (H), Asesores (I-K)
        col_inicio = {'S005': 'A', 'S010': 'E', 'S001': 'I'}
        
        # Crear títulos de todas las sociedades en la misma fila
        for sociedad_codigo in sociedades_orden:
            col_base = col_inicio[sociedad_codigo]
            nombre_sociedad = SOCIEDADES_NAMES.get(sociedad_codigo, sociedad_codigo)
            
            # Título de la sociedad (fusionar 3 columnas)
            ws.merge_cells(f'{col_base}{row}:{chr(ord(col_base) + 2)}{row}')
            ws[f'{col_base}{row}'] = nombre_sociedad
            ws[f'{col_base}{row}'].font = Font(bold=True, size=12, color="006341")
            ws[f'{col_base}{row}'].fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            ws[f'{col_base}{row}'].alignment = center_alignment
        
        row += 1
        
        # Crear encabezados de columna de todas las sociedades en la misma fila
        for sociedad_codigo in sociedades_orden:
            col_base = col_inicio[sociedad_codigo]
            ws[f'{col_base}{row}'] = 'Empresa'
            ws[f'{chr(ord(col_base) + 1)}{row}'] = 'Factura'
            ws[f'{chr(ord(col_base) + 2)}{row}'] = 'Monto (EUR)'
            
            for i in range(3):
                col = chr(ord(col_base) + i)
                cell = ws[f'{col}{row}']
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center_alignment
        
        row += 1
        
        # Escribir datos por sociedad: cada empresa con sus facturas
        # Cada sociedad se escribe de forma independiente
        fila_inicio_por_sociedad = {soc: row for soc in sociedades_orden}
        
        for sociedad_codigo in sociedades_orden:
            col_base = col_inicio[sociedad_codigo]
            empresas = datos_sociedades.get(sociedad_codigo, [])
            current_row = fila_inicio_por_sociedad[sociedad_codigo]
            
            for empresa in empresas:
                # Fila de empresa (total)
                ws[f'{col_base}{current_row}'] = f"{empresa['tercero']} - {empresa['nombre']}"
                ws[f'{chr(ord(col_base) + 1)}{current_row}'] = f"TOTAL ({len(empresa.get('facturas', []))} facturas)"
                ws[f'{chr(ord(col_base) + 2)}{current_row}'] = empresa['total']
                ws[f'{chr(ord(col_base) + 2)}{current_row}'].number_format = '#,##0.00'
                ws[f'{chr(ord(col_base) + 2)}{current_row}'].alignment = right_alignment
                
                # Estilo para fila de empresa (negrita, fondo azul claro)
                empresa_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                for i in range(3):
                    col = chr(ord(col_base) + i)
                    cell = ws[f'{col}{current_row}']
                    cell.font = Font(bold=True)
                    cell.fill = empresa_fill
                    cell.border = border
                    if i == 2:
                        cell.alignment = right_alignment
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                
                current_row += 1
                
                # Filas de facturas de esta empresa
                for factura in empresa.get('facturas', []):
                    # Usar el nombre completo de la factura si está disponible, sino tipo-asiento
                    factura_id = factura.get('nombre', '') if factura.get('nombre') else f"{factura['tipo']}-{factura['asiento']}"
                    
                    ws[f'{col_base}{current_row}'] = ''  # Vacío para indentación visual
                    ws[f'{chr(ord(col_base) + 1)}{current_row}'] = factura_id
                    ws[f'{chr(ord(col_base) + 2)}{current_row}'] = factura['monto']
                    ws[f'{chr(ord(col_base) + 2)}{current_row}'].number_format = '#,##0.00'
                    ws[f'{chr(ord(col_base) + 2)}{current_row}'].alignment = right_alignment
                    
                    # Estilo para fila de factura (fondo gris claro)
                    factura_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                    for i in range(3):
                        col = chr(ord(col_base) + i)
                        cell = ws[f'{col}{current_row}']
                        cell.border = border
                        if i == 0:
                            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                        else:
                            cell.fill = factura_fill
                        if i == 2:
                            cell.alignment = right_alignment
                        else:
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    current_row += 1
        
        # Encontrar la fila máxima para los totales
        max_row = max([fila_inicio_por_sociedad[soc] + sum(1 + len(emp.get('facturas', [])) for emp in datos_sociedades.get(soc, [])) for soc in sociedades_orden], default=row)
        row = max_row
        
        # Fila de totales (todas las sociedades en la misma fila)
        # Usar los totales validados del query del dashboard para asegurar que coincidan
        # NOTA: Las empresas se cuentan por sociedad en cada columna, pero el total general
        # debe mostrar empresas únicas (una empresa puede aparecer en múltiples sociedades)
        total_empresas_unicas = set()
        total_general = 0.0
        
        for sociedad_codigo in sociedades_orden:
            col_base = col_inicio[sociedad_codigo]
            nombre_sociedad = SOCIEDADES_NAMES.get(sociedad_codigo, sociedad_codigo)
            empresas = datos_sociedades.get(sociedad_codigo, [])
            
            # Usar el total validado del dashboard si está disponible, sino calcular
            total_sociedad = totales_validacion.get(sociedad_codigo, sum(e['total'] for e in empresas))
            total_general += total_sociedad
            
            # Contar empresas únicas (para el total general)
            for e in empresas:
                tercero_key = str(e['tercero']).strip()
                total_empresas_unicas.add(tercero_key)
            
            ws[f'{col_base}{row}'] = 'TOTAL'
            ws[f'{chr(ord(col_base) + 1)}{row}'] = f'{len(empresas)} empresas'
            ws[f'{chr(ord(col_base) + 2)}{row}'] = total_sociedad
            ws[f'{chr(ord(col_base) + 2)}{row}'].number_format = '#,##0.00'
            
            for i in range(3):
                col = chr(ord(col_base) + i)
                cell = ws[f'{col}{row}']
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                cell.border = border
                if i == 2:
                    cell.alignment = right_alignment
                else:
                    cell.alignment = center_alignment
        
        # Fila de total general (suma de las tres sociedades)
        # NOTA: El total de empresas debe ser empresas únicas (no la suma por sociedad)
        # porque una empresa puede aparecer en múltiples sociedades
        row += 1
        total_general_fill = PatternFill(start_color="006341", end_color="006341", fill_type="solid")
        total_general_font = Font(bold=True, size=14, color="FFFFFF")
        
        # Calcular total de empresas únicas (igual que el dashboard)
        total_empresas_unicas_count = len(total_empresas_unicas)
        
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = 'TOTAL GENERAL'
        ws[f'A{row}'].font = total_general_font
        ws[f'A{row}'].fill = total_general_fill
        ws[f'A{row}'].alignment = center_alignment
        
        ws.merge_cells(f'E{row}:G{row}')
        ws[f'E{row}'] = f'{total_empresas_unicas_count} empresas (únicas)'
        ws[f'E{row}'].font = total_general_font
        ws[f'E{row}'].fill = total_general_fill
        ws[f'E{row}'].alignment = center_alignment
        
        ws.merge_cells(f'I{row}:K{row}')
        ws[f'I{row}'] = total_general
        ws[f'I{row}'].number_format = '#,##0.00'
        ws[f'I{row}'].font = total_general_font
        ws[f'I{row}'].fill = total_general_fill
        ws[f'I{row}'].alignment = right_alignment
        
        # Aplicar bordes a las celdas del total general
        for col_range in [('A', 'C'), ('E', 'G'), ('I', 'K')]:
            for col in range(ord(col_range[0]), ord(col_range[1]) + 1):
                ws[f'{chr(col)}{row}'].border = border
        
        # Ajustar ancho de columnas basándose en el contenido real
        # Calcular el ancho máximo necesario para la columna Empresa (A, E, I)
        # Revisar todas las filas de datos para encontrar el nombre de empresa más largo
        max_ancho_empresa = {'A': 15, 'E': 15, 'I': 15}  # Mínimos
        
        # Revisar solo las filas de datos (desde fila_inicio hasta row-1)
        for sociedad_codigo in sociedades_orden:
            col_base = col_inicio[sociedad_codigo]
            empresas = datos_sociedades.get(sociedad_codigo, [])
            
            for empresa in empresas:
                # Calcular ancho del texto "tercero - nombre"
                texto_empresa = f"{empresa['tercero']} - {empresa['nombre']}"
                # En Excel, aproximadamente 1.2 unidades por carácter es razonable
                ancho_estimado = len(texto_empresa) * 1.2
                max_ancho_empresa[col_base] = max(max_ancho_empresa[col_base], ancho_estimado)
        
        # Aplicar anchos calculados
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            if col in ['D', 'H']:  # Columnas de espacio
                ws.column_dimensions[col].width = 3
            elif col in ['A', 'E', 'I']:  # Columna Empresa - ajustar al contenido más largo
                # Usar el máximo calculado, con límites razonables (mínimo 20, máximo 60)
                ancho = max(20, min(max_ancho_empresa.get(col, 30), 60))
                ws.column_dimensions[col].width = ancho
            elif col in ['B', 'F', 'J']:  # Columna Factura
                ws.column_dimensions[col].width = 25  # Suficiente para números de factura
            else:  # Monto (C, G, K)
                ws.column_dimensions[col].width = 18
        
        # Generar el archivo en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Preparar respuesta
        fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"informe_empresas_por_sociedad_{fecha_str}.xlsx"
        
        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generando Excel de empresas por sociedad: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Error al generar el Excel: {str(e)}")


@router.get("/api/facturas/historial-pago", response_model=dict | None, tags=["Facturas"])
def obtener_historial_pago(
    factura_id: Optional[str] = Query(None, description="Identificador de factura, p.ej. 'TYP-ASIENTO'"),
    tipo: Optional[str] = Query(None, description="Tipo de la factura (TYP_0)"),
    asiento: Optional[str] = Query(None, description="Asiento contable (ACCNUM_0)"),
    tercero: Optional[str] = Query(None, description="Código de cliente (BPR_0)"),
    sociedad: Optional[str] = Query(None, description="Código de sociedad (CPY_0)"),
    repo_facturas: RepositorioFacturas = Depends(get_repo_facturas),
):
    try:
        # Resolver factura_id
        fid = factura_id or (f"{tipo}-{asiento}" if (tipo and asiento) else None)
        if not fid:
            return None

        # Leer pago desde la tabla de SAGE X3 (x3v12.ATISAINT.facturas_cambio_pago)
        from sqlalchemy import text as sqltext
        try:
            q_pago = sqltext(
                """
                SELECT TOP 1 factura_id, fecha_cambio, monto_pagado, BPR_0
                FROM ATISAINT.facturas_cambio_pago
                WHERE factura_id = :fid
                ORDER BY fecha_cambio DESC
                """
            )
            row_pago = repo_facturas.db.execute(q_pago, {"fid": fid}).mappings().first()
        except Exception:
            row_pago = None
        if not row_pago:
            return None

        # Intentar enriquecer con vencimiento desde la BD de facturas reales
        try:
            partes = str(fid).split('-')
            if len(partes) >= 2:
                q = sqltext("""
                    SELECT TOP 1 DUDDAT_0 AS venc
                    FROM x3v12.ATISAINT.GACCDUDATE
                    WHERE TYP_0 = :tipo AND ACCNUM_0 = :asiento
                """)
                row = repo_facturas.db.execute(q, {"tipo": partes[0], "asiento": partes[1]}).first()
                if row and hasattr(row, 'venc'):
                    venc = row.venc
                else:
                    venc = None
            else:
                venc = None
        except Exception:
            venc = None

        # Calcular días de retraso si es posible
        dias = None
        try:
            if venc is not None and row_pago.get('fecha_cambio') is not None:
                from datetime import date
                v = row_pago['fecha_cambio']
                # v y venc pueden ser datetime/date; convertir a date
                fp = v.date() if hasattr(v, 'date') else v
                vc = venc.date() if hasattr(venc, 'date') else venc
                dias = (fp - vc).days
                if dias < 0:
                    dias = 0
        except Exception:
            dias = None

        return {
            "factura_id": fid,
            "fecha_pago": row_pago.get('fecha_cambio'),
            "vencimiento": venc,
            "dias_retraso": dias,
            "monto_pagado": row_pago.get('monto_pagado'),
        }
    except Exception as e:
        logger.error(f"Error al obtener historial de pago: {e}")
        # Preferimos no romper el frontend: devolver None
        return None


@router.get("/api/facturas/historial-pagadas", response_model=List[dict], tags=["Facturas"])
def listar_historial_pagadas(
    tercero: Optional[str] = Query(None, description="Código de cliente (BPR_0) para filtrar"),
    factura_id: Optional[str] = Query(None, description="ID específico de factura (ROWID) para filtrar"),
    limit: int = Query(200, ge=1, le=1000),
    repo_gestion: RepositorioGestion = Depends(get_repo_gestion),
    repo_facturas: RepositorioFacturas = Depends(get_repo_facturas),
):
    """Lista facturas pagadas (vencidas) a partir de la tabla externa `facturas_cambio_pago`.

    Nota: `facturas_cambio_pago` no guarda el tercero; se resuelve por `tipo` y `asiento`
    consultando la tabla real de facturas. Si `tercero` se proporciona, se filtra por él.
    """
    try:
        from sqlalchemy import text as sqltext
        
        # Verificar si la tabla existe primero
        try:
            check_table_sql = "SELECT COUNT(*) as count FROM ATISAINT.facturas_cambio_pago"
            result = repo_facturas.db.execute(sqltext(check_table_sql)).first()
        except Exception as e:
            logger.error(f"Error verificando tabla facturas_cambio_pago: {e}")
            return []
        
        # Si no hay datos, devolver array vacío
        if not result or result[0] == 0:
            return []
        
        # Construir SQL según dialecto
        try:
            dialect = repo_facturas.db.bind.dialect.name if hasattr(repo_facturas.db, 'bind') else 'mssql'
        except Exception:
            dialect = 'mssql'

        # Leer desde la BD de facturas (SAGE X3): esquema ATISAINT
        table_name = 'ATISAINT.facturas_cambio_pago'
        # MSSQL: TOP, otros: LIMIT
      
        base_sql = f"SELECT factura_id, fecha_cambio, monto_pagado, idcliente FROM {table_name}"

        filtros = []
        params = {}
        if tercero:
            filtros.append("idcliente = :tercero")
            params["tercero"] = str(tercero)
        if factura_id:
            filtros.append("factura_id = :factura_id")
            params["factura_id"] = str(factura_id)

        where_sql = (" WHERE " + " AND ".join(filtros)) if filtros else ""
        order_sql = " ORDER BY fecha_cambio DESC"
        tail_sql = "" if dialect == 'mssql' else f" LIMIT {int(limit)}"

        q = sqltext(base_sql + where_sql + order_sql + tail_sql)
        rows = repo_facturas.db.execute(q, params).mappings().all()

        def _norm(v: Optional[str]) -> Optional[str]:
            if v is None:
                return None
            s = str(v).strip()
            if s == "":
                return s
            # Normalizar ceros a la izquierda si es numérico
            try:
                # Mantener no numéricos tal cual
                if s.isdigit():
                    return str(int(s))
            except Exception:
                pass
            return s

        tercero_norm = _norm(tercero) if tercero else None

        # Agrupar pagos por factura
        facturas_pagos = {}
        for r in rows:
            factura_id = str(r.get('factura_id'))
            if not factura_id:
                continue
                
            if factura_id not in facturas_pagos:
                facturas_pagos[factura_id] = {
                    'pagos': [],
                    'total_pagado': 0,
                    'datos_factura': None
                }
            
            # Agregar pago a la factura (puede haber múltiples pagos para la misma factura)
            monto_pagado = r.get('monto_pagado') or 0
            fecha_pago = r.get('fecha_cambio')
            
            # Verificar si ya existe un pago con la misma fecha para evitar duplicados
            pago_existente = False
            for pago in facturas_pagos[factura_id]['pagos']:
                if pago['fecha_pago'] == fecha_pago and pago['monto_pagado'] == monto_pagado:
                    pago_existente = True
                    break
            
            if not pago_existente:
                facturas_pagos[factura_id]['pagos'].append({
                    'fecha_pago': fecha_pago,
                    'monto_pagado': monto_pagado
                })
                facturas_pagos[factura_id]['total_pagado'] += monto_pagado
                
                # Ordenar pagos por fecha (más reciente primero)
                facturas_pagos[factura_id]['pagos'].sort(
                    key=lambda x: x['fecha_pago'] if x['fecha_pago'] else '', 
                    reverse=True
                )
            
            # Buscar datos de la factura solo una vez
            if facturas_pagos[factura_id]['datos_factura'] is None:
                try:
                    gaccdudate_sql = """
                        SELECT TOP 1 BPR_0 AS tercero, DUDDAT_0 AS vencimiento, CPY_0 AS sociedad, 
                               TYP_0 AS tipo, ACCNUM_0 AS asiento, ROWID, AMTCUR_0 AS importe_total
                        FROM x3v12.ATISAINT.GACCDUDATE
                        WHERE ROWID = :factura_id
                    """
                    gaccdudate_result = repo_facturas.db.execute(sqltext(gaccdudate_sql), {"factura_id": factura_id}).mappings().first()
                    
                    if gaccdudate_result:
                        facturas_pagos[factura_id]['datos_factura'] = gaccdudate_result
                    else:
                        # Datos básicos si no se encuentra en GACCDUDATE
                        facturas_pagos[factura_id]['datos_factura'] = {
                            'tercero': r.get('idcliente'),
                            'vencimiento': None,
                            'sociedad': None,
                            'tipo': None,
                            'asiento': None,
                            'importe_total': None
                        }
                except Exception as e:
                    logger.warning(f"Error consultando GACCDUDATE para factura {factura_id}: {e}")
                    facturas_pagos[factura_id]['datos_factura'] = {
                        'tercero': r.get('idcliente'),
                        'vencimiento': None,
                        'sociedad': None,
                        'tipo': None,
                        'asiento': None,
                        'importe_total': None
                    }
        
        # Convertir a formato de salida
        out: List[dict] = []
        for factura_id, datos in facturas_pagos.items():
            datos_factura = datos['datos_factura']
            
            # Aplicar filtro por tercero si se especifica
            tercero_factura = datos_factura.get('tercero')
            if tercero and tercero_norm and _norm(tercero_factura) != tercero_norm:
                continue
            
            # Calcular días de retraso (usar el primer pago - el más temprano)
            dias_retraso = None
            if datos['pagos'] and datos_factura.get('vencimiento'):
                try:
                    venc = datos_factura['vencimiento']
                    # Usar el último pago de la lista ordenada (el más temprano)
                    primer_pago = datos['pagos'][-1]['fecha_pago']
                    
                    # Convertir fechas a objetos date
                    if hasattr(venc, 'date'):
                        vc = venc.date()
                    elif isinstance(venc, str):
                        from datetime import datetime
                        vc = datetime.fromisoformat(venc.replace('Z', '+00:00')).date()
                    else:
                        vc = venc
                    
                    if hasattr(primer_pago, 'date'):
                        fp = primer_pago.date()
                    elif isinstance(primer_pago, str):
                        from datetime import datetime
                        fp = datetime.fromisoformat(primer_pago.replace('Z', '+00:00')).date()
                    else:
                        fp = primer_pago
                    
                    dias_retraso = (fp - vc).days
                    if dias_retraso < 0:
                        dias_retraso = 0
                except Exception as e:
                    logger.warning(f"Error calculando días de retraso: {e}")
                    dias_retraso = None
            
            # Calcular importe pendiente
            importe_total = datos_factura.get('importe_total') or 0
            importe_pendiente = max(0, importe_total - datos['total_pagado'])
            
            out.append({
                "factura_id": factura_id,
                "tipo": datos_factura.get('tipo'),
                "asiento": datos_factura.get('asiento'),
                "tercero": tercero_factura,
                "sociedad": datos_factura.get('sociedad'),
                "vencimiento": datos_factura.get('vencimiento'),
                "importe_total": importe_total,
                "total_pagado": datos['total_pagado'],
                "importe_pendiente": importe_pendiente,
                "dias_retraso": dias_retraso,
                "pagos": datos['pagos']
            })

        return out
    except Exception as e:
        logger.error(f"Error al listar historial pagadas: {e}")
        return []


@router.get("/api/facturas/historial-pagadas/debug", tags=["Facturas"])
def debug_historial_pagadas(
    repo_facturas: RepositorioFacturas = Depends(get_repo_facturas),
):
    """Endpoint de diagnóstico para ver la estructura de la tabla facturas_cambio_pago."""
    try:
        from sqlalchemy import text as sqltext
        
        # Verificar estructura de la tabla
        try:
            # Intentar leer las primeras filas para ver qué columnas existen
            debug_sql = "SELECT TOP 1 * FROM ATISAINT.facturas_cambio_pago"
            result = repo_facturas.db.execute(sqltext(debug_sql)).mappings().first()
            
            if result:
                return {
                    "message": "Estructura de tabla encontrada",
                    "columns": list(result.keys()),
                    "sample_data": dict(result)
                }
            else:
                return {"message": "Tabla vacía o no existe"}
                
        except Exception as e:
            return {"error": f"Error consultando tabla: {str(e)}"}
            
    except Exception as e:
        return {"error": str(e)}


def get_repo_registro(db: Session = Depends(get_gestion_db)):
    return RepositorioRegistroFacturas(db)


